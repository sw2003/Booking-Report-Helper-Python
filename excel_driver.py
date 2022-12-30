import openpyxl
import ship_class
import customtkinter

def driver(filepaths, progress_label, progress_bar):

    report_filepath = filepaths["report_filepath"] 
    data_filepath = filepaths["data_filepath"]

    read_workbook = openpyxl.load_workbook(data_filepath)
    read_worksheet = read_workbook.active

    write_workbook = openpyxl.load_workbook(report_filepath)
    write_worksheet = write_workbook.active

    start_index = 5
    current_shipping_lane = 'CPNW'

    def end_index():
        i = (write_worksheet.max_row)//2

        row_val = write_worksheet['A' + str(i)].value

        if (row_val != None):
            row_val = row_val.strip()
        
        while row_val != "SUB TOTAL":  
            i = i + 1
            row_val = write_worksheet['A' + str(i)].value

            if (row_val != None):
                row_val = row_val.strip()

            if (i == 150):
                return -1
        return i


    for i in range(start_index, end_index()):
        progress_label.configure(text = str(i) + f"/{end_index()-1}")
        progress_bar.set(i/end_index()) 
        print(i)
        valid_shipping_lanes = ['CPNW', 'CENX', 'MPNW', 'OPNW', 'EPNW', 'AWE5', 'GEX1', 'GEX2','CEN ( extra loader)', 'CEN']
        invalid_keywords = ['TBA', 'TBN', 'AAC ( extra loader)', 'BLANK VOYAGE']
        directions = ['N', 'W', 'S', 'E'] 

        value = write_worksheet['A' + str(i)].value

        #print(dir(worksheet['A' + str(i)])) 

        if (value != None):
            value = value.strip()

        if value in valid_shipping_lanes:
            current_shipping_lane = value

            #if value == 'CEN ( extra loader)':
            #  current_shipping_lane = 'CEN' 

        ### Validate if row contains valid shipping data

        three_digit_code = None
        vessel_code = None
        
        if value != None and value not in invalid_keywords: 
            value = value.strip()

            if write_worksheet['B'+str(i)] != None: 
                
                if value not in valid_shipping_lanes:
                    arr = value.split(" ")
                    if (arr[len(arr)-1] in directions): 
                        arr.pop()
                    
                    three_digit_code = arr[len(arr)-1] 
                    vessel_code = write_worksheet['B'+str(i)].value

                    ship = ship_class.ship(current_shipping_lane,vessel_code,three_digit_code, i, read_worksheet, write_worksheet, write_workbook, report_filepath)
                    ship.create_name() 
                    ship.get_data() 
                    ship.paste_data() 
                    ship.save_data() 
                    progress_label.configure(text = "Complete")

                



    
    
